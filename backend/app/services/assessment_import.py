"""Parse, match and diff uploaded skill assessments.

Pure functions over bytes and already-loaded ORM rows: nothing here touches HTTP or
commits a transaction, so the whole matching/diffing surface is testable without a
database. `assessment_apply.py` owns the write.

Two things about the shipped workbooks drive the design here, both measured rather
than assumed (openpyxl 3.1.5):

1. The derived columns on the Assessment sheet cache as empty. In the fully populated
   EXAMPLE workbook, `Assessment!K5` (Gap) reads None. In the TEMPLATE — the file a
   user actually fills in — `Assessment!A5` (Employee) is itself a formula
   (`IF(Roster!$A$5="","",Roster!$A$5)`) that reads None. Importing the Assessment
   sheet's derived columns from a file that was never recalculated in Excel would
   match zero rows and report a clean "0 changes".

   So identity and self ratings are read from the *literal* source sheets — `Roster`
   and `Intake` — and only the genuinely literal Assessment columns (reviewer rating,
   reviewer evidence, reviewed-on) are read from `Assessment`.

2. `read_only=True` raises AttributeError on `cell.hyperlink`, and evidence is written
   as `{text, hyperlink}`, so the workbook must be opened read-write to recover
   `evidence_url`.
"""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Iterable, Literal, Optional

import openpyxl

# Rating anchors, mirroring PROFICIENCY_LABELS in lib/types.ts and the legacy
# four-adjective scale that lib/utils.ts:parseProficiency still accepts.
PROFICIENCY_LABELS = [
    "not exposed",
    "aware",
    "guided practitioner",
    "independent",
    "advanced/lead",
    "strategic expert",
]

LEGACY_LEVELS = {
    "none": 0, "not exposed": 0,
    "beginner": 1, "novice": 1, "aware": 1,
    "intermediate": 2, "basic": 2, "guided": 2, "guided practitioner": 2,
    "advanced": 3, "independent": 3, "proficient": 3,
    "expert": 4, "lead": 4, "advanced/lead": 4,
    "strategic": 5, "strategic expert": 5,
}

# Cell text that explicitly clears a stored value. A *blank* cell means "leave
# unchanged" (matching exclude_unset= on the existing single-cell POST), so there has
# to be some way to say "actually, remove this".
CLEAR_TOKENS = {"-", "--", "none", "clear"}

MAX_UPLOAD_BYTES = 10 * 1024 * 1024
MAX_ROWS = 50_000

RowStatus = Literal[
    "ok",
    "unchanged",
    "empty",
    "duplicate",
    "unknown_employee",
    "ambiguous_employee",
    "unknown_skill",
    "invalid_value",
    "forbidden_field",
]

WRITABLE_FIELDS = ("target_override", "self_rating", "reviewer_rating", "evidence", "evidence_url")
RATING_FIELDS = ("target_override", "self_rating", "reviewer_rating")

MANAGER_FIELDS = frozenset(WRITABLE_FIELDS)
SELF_FIELDS = frozenset({"self_rating", "evidence"})


class ImportError_(Exception):
    """Raised for a file that cannot be parsed at all (wrong type, corrupt, unrecalculated)."""

    def __init__(self, message: str, status_code: int = 400):
        super().__init__(message)
        self.message = message
        self.status_code = status_code


# ---------------------------------------------------------------------------
# Header normalization and column matching
# ---------------------------------------------------------------------------

def normalize_header(label: Any) -> str:
    """Fold a header cell to a comparable key.

    Turns `MY EVIDENCE — what would you point at?` into `my evidence` and
    `Key (calculated)` into `key`, so the alias tables below stay readable.
    """
    text = "" if label is None else str(label)
    text = text.split("—")[0].split(" - ")[0].split(":")[0]
    text = re.sub(r"\([^)]*\)", " ", text)
    text = text.replace("_", " ").replace("-", " ").replace("/", " ")
    text = text.strip().rstrip("?").lower()
    return re.sub(r"\s+", " ", text).strip()


# Logical field -> aliases. Resolution is exact-match-first across all columns, then a
# keyword fallback; that ordering is what stops `Employee ID` being claimed by the
# `employee` alias and `Skill ID` by `skill`.
COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "employee_email": ("email", "e mail", "email address", "work email", "employee email", "upn"),
    "employee_code": ("employee id", "employee number", "emp id", "staff id", "badge", "worker id"),
    "employee_name": (
        "employee", "name", "employee name", "engineer", "person", "team member",
        "full name", "display name", "assessee",
    ),
    "skill_code": ("skill id", "skill code", "skill number", "code", "id"),
    "skill_name": ("skill", "skill name", "capability", "competency"),
    "target": ("target", "target level", "target rating", "team target", "expected level"),
    "self_rating": ("self", "self rating", "self assessment", "self level", "self score", "my rating"),
    "reviewer_rating": (
        "reviewer", "reviewer rating", "reviewer level", "reviewer score",
        "manager rating", "lead rating", "calibrated rating",
    ),
    "self_evidence": ("self evidence", "my evidence"),
    "evidence": ("evidence", "reviewer evidence", "notes", "justification", "examples"),
    "evidence_url": ("evidence url", "url", "link"),
    "assessed_at": ("assessment date", "assessed on", "reviewed on", "submitted", "date", "as of"),
    # Read for warnings only — never written. See module docstring.
    "file_final": ("final", "final rating", "final level"),
    "file_gap": ("gap", "gap to target"),
    "file_priority": ("priority", "gap priority"),
}

# Fallback keyword sets, used only when no exact alias hit. Every keyword must appear.
# `manager` deliberately requires a rating word: the app's own export writes the
# manager's *name* in a column headed `Manager` (lib/skill-workbook.ts:109).
COLUMN_KEYWORDS: dict[str, tuple[tuple[str, ...], ...]] = {
    "employee_email": (("email",),),
    "employee_code": (("employee", "id"), ("staff", "id")),
    "employee_name": (("employee",), ("engineer",)),
    "skill_code": (("skill", "id"), ("skill", "code")),
    "skill_name": (("skill",),),
    "target": (("target",),),
    "self_rating": (("self", "rating"), ("my", "rating")),
    "reviewer_rating": (("reviewer", "rating"), ("manager", "rating"), ("reviewer",)),
    "evidence": (("evidence",),),
    "assessed_at": (("date",),),
}


def match_columns(headers: Iterable[Any]) -> dict[str, int]:
    """Map logical field -> 0-based column index for one header row."""
    normalized = [normalize_header(h) for h in headers]
    taken: set[int] = set()
    resolved: dict[str, int] = {}

    # Pass 1: exact alias equality, most-specific fields first so `employee id` and
    # `skill id` claim their columns before the looser `employee` / `skill` aliases.
    order = (
        "employee_email", "employee_code", "skill_code", "self_evidence", "evidence_url",
        "target", "self_rating", "reviewer_rating", "file_final", "file_gap", "file_priority",
        "assessed_at", "evidence", "employee_name", "skill_name",
    )
    for field_name in order:
        for idx, label in enumerate(normalized):
            if idx in taken or not label:
                continue
            if label in COLUMN_ALIASES[field_name]:
                resolved[field_name] = idx
                taken.add(idx)
                break

    # Pass 2: keyword fallback for anything still unbound.
    for field_name in order:
        if field_name in resolved or field_name not in COLUMN_KEYWORDS:
            continue
        for keywords in COLUMN_KEYWORDS[field_name]:
            hit = next(
                (
                    idx for idx, label in enumerate(normalized)
                    if idx not in taken and label and all(k in label for k in keywords)
                ),
                None,
            )
            if hit is not None:
                resolved[field_name] = hit
                taken.add(hit)
                break

    return resolved


def find_header_row(rows: list[list[Any]], anchors: tuple[str, ...], limit: int = 15) -> Optional[int]:
    """First row (0-based) whose cells contain every anchor as a substring.

    Locating the header by content rather than by a fixed position is what lets both
    workbook generators — which put the same logical columns in different places — and
    a hand-built sheet all be read by the same code.
    """
    for i, row in enumerate(rows[:limit]):
        labels = [normalize_header(c) for c in row]
        if all(any(a in label for label in labels if label) for a in anchors):
            return i
    return None


# ---------------------------------------------------------------------------
# Value parsing
# ---------------------------------------------------------------------------

def parse_level(value: Any) -> tuple[Optional[int], bool, Optional[str]]:
    """Parse a 0-5 rating.

    Returns (level, is_clear, error). Deliberately does NOT clamp: lib/utils.ts
    clampLevel() silently turns a typo'd 45 into a 5, which is exactly the sort of
    thing an import should refuse rather than quietly write.
    """
    if value is None:
        return None, False, None
    if isinstance(value, bool):
        return None, False, f"{value!r} is not a rating"

    if isinstance(value, (int, float)):
        if float(value) != int(value):
            return None, False, f"{value} is not a whole number"
        level = int(value)
        if not 0 <= level <= 5:
            return None, False, f"{level} is outside the 0-5 scale"
        return level, False, None

    text = str(value).strip()
    if not text:
        return None, False, None
    if text.lower() in CLEAR_TOKENS:
        return None, True, None

    if re.fullmatch(r"-?\d+(\.0+)?", text):
        level = int(float(text))
        if not 0 <= level <= 5:
            return None, False, f"{text} is outside the 0-5 scale"
        return level, False, None

    key = text.lower()
    if key in PROFICIENCY_LABELS:
        return PROFICIENCY_LABELS.index(key), False, None
    if key in LEGACY_LEVELS:
        return LEGACY_LEVELS[key], False, None
    return None, False, f"{text!r} is not a rating"


def parse_text(value: Any) -> tuple[Optional[str], bool]:
    """Returns (text, is_clear). Blank means 'leave unchanged'."""
    if value is None:
        return None, False
    text = str(value).strip()
    if not text:
        return None, False
    if text.lower() in CLEAR_TOKENS:
        return None, True
    return text, False


def parse_date(value: Any) -> Optional[datetime]:
    if isinstance(value, datetime):
        return value
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:10], fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


@dataclass
class ParsedRow:
    """One source row, before it is matched against the database."""
    row_number: int
    # A list of pairs, not a dict: lib/api-client.ts transformKeys() camelCases every
    # key at every depth, so dict keys taken from a user's own CSV headers would come
    # back rewritten into something that no longer matches their file.
    raw: list[list[str]] = field(default_factory=list)
    employee_email: Optional[str] = None
    employee_code: Optional[str] = None
    employee_name: Optional[str] = None
    skill_code: Optional[str] = None
    skill_name: Optional[str] = None
    values: dict[str, Any] = field(default_factory=dict)   # writable fields present in the row
    clears: set[str] = field(default_factory=set)          # fields explicitly set to NULL
    file_derived: dict[str, Any] = field(default_factory=dict)  # Final/Gap/Priority, warnings only
    assessed_at: Optional[datetime] = None
    errors: list[str] = field(default_factory=list)

    def has_identity(self) -> bool:
        return bool(self.employee_email or self.employee_code or self.employee_name)


def _row_from_cells(
    row_number: int,
    cells: list[Any],
    headers: list[str],
    cols: dict[str, int],
    *,
    hyperlinks: Optional[dict[int, str]] = None,
) -> ParsedRow:
    """Build a ParsedRow from one sheet/CSV row given a resolved column map."""
    def cell(name: str) -> Any:
        idx = cols.get(name)
        if idx is None or idx >= len(cells):
            return None
        return cells[idx]

    parsed = ParsedRow(row_number=row_number)
    parsed.raw = [
        [headers[i] if i < len(headers) else f"col{i + 1}", "" if v is None else str(v)]
        for i, v in enumerate(cells)
        if v is not None and str(v).strip() != ""
    ]

    parsed.employee_email, _ = parse_text(cell("employee_email"))
    parsed.employee_code, _ = parse_text(cell("employee_code"))
    parsed.employee_name, _ = parse_text(cell("employee_name"))
    parsed.skill_code, _ = parse_text(cell("skill_code"))
    parsed.skill_name, _ = parse_text(cell("skill_name"))
    parsed.assessed_at = parse_date(cell("assessed_at"))

    for source, target in (("target", "target_override"), ("self_rating", "self_rating"), ("reviewer_rating", "reviewer_rating")):
        if source not in cols:
            continue
        level, is_clear, error = parse_level(cell(source))
        if error:
            parsed.errors.append(f"{source}: {error}")
        elif is_clear:
            parsed.clears.add(target)
        elif level is not None:
            parsed.values[target] = level

    # A sheet may carry reviewer evidence and self evidence in separate columns; the
    # reviewer's wins when both are present, since it is the calibrated statement.
    for source in ("self_evidence", "evidence"):
        if source not in cols:
            continue
        text, is_clear = parse_text(cell(source))
        if is_clear:
            parsed.clears.add("evidence")
            parsed.values.pop("evidence", None)
        elif text is not None:
            parsed.values["evidence"] = text
            parsed.clears.discard("evidence")

    url, url_clear = parse_text(cell("evidence_url"))
    if url_clear:
        parsed.clears.add("evidence_url")
    elif url:
        parsed.values["evidence_url"] = url
    elif hyperlinks:
        # buildWorkbook writes evidence as {text, hyperlink}, so when there is no
        # separate URL column the link lives on the evidence cell itself.
        for name in ("evidence", "self_evidence"):
            idx = cols.get(name)
            if idx is not None and idx in hyperlinks:
                parsed.values["evidence_url"] = hyperlinks[idx]
                break

    for name in ("file_final", "file_gap", "file_priority"):
        value = cell(name)
        if value is not None and str(value).strip() != "":
            parsed.file_derived[name] = str(value).strip()

    return parsed


def _sheet_rows(ws) -> list[list[Any]]:
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _hyperlinks_for_row(ws, excel_row: int) -> dict[int, str]:
    links: dict[int, str] = {}
    for cell in ws[excel_row]:
        target = getattr(cell.hyperlink, "target", None) if cell.hyperlink else None
        if target:
            links[cell.column - 1] = target
    return links


ASSESSMENT_ANCHORS = ("employee", "skill id")
INTAKE_ANCHORS = ("employee", "skill id", "self rating")
MY_ASSESSMENT_ANCHORS = ("skill id", "my rating")


def _find_sheet(wb, *keywords: str):
    for ws in wb.worksheets:
        name = ws.title.strip().lower()
        if all(k in name for k in keywords):
            return ws
    return None


def parse_csv_bytes(data: bytes) -> tuple[list[ParsedRow], list[str]]:
    text = data.decode("utf-8-sig", errors="replace")
    dialect_delim = "\t" if text.count("\t") > text.count(",") else ","
    reader = csv.reader(io.StringIO(text), delimiter=dialect_delim)
    all_rows = [r for r in reader]
    if not all_rows:
        raise ImportError_("That file is empty.")

    header_index = find_header_row(all_rows, ("skill",)) or 0
    headers = [normalize_header(h) for h in all_rows[header_index]]
    cols = match_columns(all_rows[header_index])
    warnings: list[str] = []

    if not any(k in cols for k in ("employee_email", "employee_code", "employee_name")):
        raise ImportError_(
            "No employee column found. Add an Email, Employee ID or Employee column."
        )
    if not any(k in cols for k in ("skill_code", "skill_name")):
        raise ImportError_("No skill column found. Add a Skill ID or Skill column.")

    rows: list[ParsedRow] = []
    for offset, cells in enumerate(all_rows[header_index + 1:], start=header_index + 2):
        if all(c is None or str(c).strip() == "" for c in cells):
            continue
        rows.append(_row_from_cells(offset, cells, headers, cols))
        if len(rows) > MAX_ROWS:
            raise ImportError_(f"That file has more than {MAX_ROWS:,} rows.")
    return rows, warnings


def _merge_key(name: Optional[str], code: Optional[str]) -> Optional[str]:
    if not name or not code:
        return None
    return f"{name.strip().lower()}|{str(code).strip()}"


def parse_xlsx_bytes(data: bytes) -> tuple[list[ParsedRow], list[str]]:
    """Read an .xlsx, preferring literal source sheets over derived ones.

    See the module docstring: on the manager workbook the Assessment sheet's Employee
    and Self rating columns are formulas whose cached values can be empty, so identity
    and self ratings come from `Roster`/`Intake` and only reviewer-side columns are
    read from `Assessment`.
    """
    try:
        # read_only=False is required: ReadOnlyCell has no .hyperlink, and that is
        # where evidence_url comes from.
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception as exc:  # zipfile.BadZipFile, KeyError on a malformed package...
        raise ImportError_(f"That file could not be read as a spreadsheet ({type(exc).__name__}).")

    warnings: list[str] = []
    rows: list[ParsedRow] = []

    intake_ws = _find_sheet(wb, "intake")
    assessment_ws = _find_sheet(wb, "assessment") or _find_sheet(wb, "my assessment")
    my_assessment_ws = _find_sheet(wb, "my", "assessment")

    # --- engineer-facing self-assessment workbook -------------------------------
    # Previously rejected outright (lib/skill-workbook.ts:563). Its `My Assessment`
    # sheet holds the engineer's literal input cells; `Send to Manager` is all
    # formulas and is never parsed.
    if my_assessment_ws is not None and intake_ws is None:
        sheet_rows = _sheet_rows(my_assessment_ws)
        header_index = find_header_row(sheet_rows, MY_ASSESSMENT_ANCHORS)
        if header_index is not None:
            cols = match_columns(sheet_rows[header_index])
            headers = [normalize_header(h) for h in sheet_rows[header_index]]
            name = _self_assessment_name(wb)
            for offset, cells in enumerate(sheet_rows[header_index + 1:], start=header_index + 2):
                if all(c is None or str(c).strip() == "" for c in cells):
                    continue
                row = _row_from_cells(offset, cells, headers, cols)
                if row.employee_name is None:
                    row.employee_name = name
                if row.values or row.clears:
                    rows.append(row)
            if name is None:
                warnings.append(
                    "This self-assessment workbook has no name on its 'Start Here' sheet, "
                    "so its rows cannot be matched to a person."
                )
            return rows, warnings

    # --- manager workbook -------------------------------------------------------
    by_key: dict[str, ParsedRow] = {}

    if intake_ws is not None:
        sheet_rows = _sheet_rows(intake_ws)
        header_index = find_header_row(sheet_rows, INTAKE_ANCHORS)
        if header_index is not None:
            cols = match_columns(sheet_rows[header_index])
            headers = [normalize_header(h) for h in sheet_rows[header_index]]
            for offset, cells in enumerate(sheet_rows[header_index + 1:], start=header_index + 2):
                if all(c is None or str(c).strip() == "" for c in cells):
                    continue
                row = _row_from_cells(offset, cells, headers, cols)
                if not (row.values or row.clears or row.errors):
                    continue
                rows.append(row)
                key = _merge_key(row.employee_name, row.skill_code)
                if key:
                    by_key[key] = row

    if assessment_ws is not None and assessment_ws is not my_assessment_ws:
        sheet_rows = _sheet_rows(assessment_ws)
        header_index = find_header_row(sheet_rows, ASSESSMENT_ANCHORS)
        if header_index is not None:
            cols = match_columns(sheet_rows[header_index])
            headers = [normalize_header(h) for h in sheet_rows[header_index]]
            data_rows = sheet_rows[header_index + 1:]

            identity_seen = any(
                _row_identity_present(cells, cols) for cells in data_rows
            )
            if not identity_seen and data_rows:
                raise ImportError_(
                    "This workbook's Employee column is a formula with no saved values, so no "
                    "row can be matched to a person. Open it in Excel and save it (which "
                    "recalculates the sheet), or fill in the Intake sheet and import that."
                )

            for offset, cells in enumerate(data_rows, start=header_index + 2):
                if all(c is None or str(c).strip() == "" for c in cells):
                    continue
                links = _hyperlinks_for_row(assessment_ws, offset)
                row = _row_from_cells(offset, cells, headers, cols, hyperlinks=links)
                # Self rating on this sheet is an INDEX/MATCH into Intake; the Intake
                # sheet is the literal source and already contributed it.
                key = _merge_key(row.employee_name, row.skill_code)
                existing = by_key.get(key) if key else None
                if existing is not None:
                    for name in ("reviewer_rating", "evidence", "evidence_url"):
                        if name in row.values:
                            existing.values[name] = row.values[name]
                    existing.clears |= {c for c in row.clears if c != "self_rating"}
                    existing.file_derived.update(row.file_derived)
                    existing.errors.extend(row.errors)
                    if row.assessed_at and not existing.assessed_at:
                        existing.assessed_at = row.assessed_at
                    continue
                row.values.pop("self_rating", None)
                if row.values or row.clears or row.errors:
                    rows.append(row)
                    if key:
                        by_key[key] = row

    if not rows:
        raise ImportError_(
            "No assessment rows found. Expected an 'Assessment', 'Intake' or "
            "'My Assessment' sheet with Employee and Skill ID columns."
        )
    if len(rows) > MAX_ROWS:
        raise ImportError_(f"That file has more than {MAX_ROWS:,} rows.")
    return rows, warnings


def _row_identity_present(cells: list[Any], cols: dict[str, int]) -> bool:
    for name in ("employee_email", "employee_code", "employee_name"):
        idx = cols.get(name)
        if idx is None or idx >= len(cells):
            continue
        if cells[idx] is not None and str(cells[idx]).strip():
            return True
    return False


def _self_assessment_name(wb) -> Optional[str]:
    """The engineer's name, entered on the 'Start Here' sheet (cell B4)."""
    ws = _find_sheet(wb, "start")
    if ws is None:
        return None
    value = ws["B4"].value
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.startswith("("):
        return None
    return text


def parse_bytes(filename: str, data: bytes) -> tuple[list[ParsedRow], list[str], str]:
    """Dispatch on extension. Returns (rows, warnings, source)."""
    if len(data) > MAX_UPLOAD_BYTES:
        raise ImportError_(
            f"That file is larger than {MAX_UPLOAD_BYTES // (1024 * 1024)} MB.", status_code=413
        )
    lower = (filename or "").lower()
    if lower.endswith((".csv", ".tsv", ".txt")):
        rows, warnings = parse_csv_bytes(data)
        return rows, warnings, "csv"
    if lower.endswith((".xlsx", ".xlsm")):
        rows, warnings = parse_xlsx_bytes(data)
        return rows, warnings, "xlsx"
    raise ImportError_(
        "Unsupported file type. Upload a .csv, .tsv, .xlsx or .xlsm file "
        "(legacy .xls is not supported).",
        status_code=415,
    )


# ---------------------------------------------------------------------------
# Matching and diffing
# ---------------------------------------------------------------------------

@dataclass
class ResolvedRow:
    row_number: int
    status: RowStatus
    employee_id: Optional[str] = None
    employee_name: Optional[str] = None
    skill_id: Optional[str] = None
    skill_name: Optional[str] = None
    matched_by: Optional[str] = None  # "email" | "employee_id" | "name"
    values: dict[str, Any] = field(default_factory=dict)
    before: dict[str, Any] = field(default_factory=dict)
    messages: list[str] = field(default_factory=list)
    raw: list[list[str]] = field(default_factory=list)
    assessed_at: Optional[str] = None

    def to_json(self) -> dict[str, Any]:
        return {
            "row_number": self.row_number,
            "status": self.status,
            "employee_id": self.employee_id,
            "employee_name": self.employee_name,
            "skill_id": self.skill_id,
            "skill_name": self.skill_name,
            "matched_by": self.matched_by,
            "values": self.values,
            "before": self.before,
            "messages": self.messages,
            "raw": self.raw,
            "assessed_at": self.assessed_at,
        }


COUNT_KEYS = (
    "rows_read", "ok", "unchanged", "empty", "duplicate", "unknown_employee",
    "ambiguous_employee", "unknown_skill", "invalid_value", "forbidden_field",
)


@dataclass
class ImportPreviewResult:
    rows: list[ResolvedRow]
    counts: dict[str, int]
    warnings: list[str]


def _index_employees(employees) -> tuple[dict, dict, dict]:
    """Three lookup maps. A key hitting more than one person maps to None (ambiguous)."""
    by_email: dict[str, Any] = {}
    by_code: dict[str, Any] = {}
    by_name: dict[str, Any] = {}

    def put(mapping: dict, key: Optional[str], employee) -> None:
        if not key:
            return
        mapping[key] = None if key in mapping else employee

    for employee in employees:
        put(by_email, (employee.email or "").strip().lower() or None, employee)
        put(by_code, (employee.employee_id or "").strip().upper() or None, employee)
        put(by_name, re.sub(r"\s+", " ", (employee.name or "").strip().lower()) or None, employee)
    return by_email, by_code, by_name


def resolve_rows(
    parsed_rows: list[ParsedRow],
    employees,
    catalog,
    existing,
    *,
    allowed_fields: frozenset[str] = MANAGER_FIELDS,
    restrict_to_employee_id: Optional[str] = None,
) -> ImportPreviewResult:
    """Match each parsed row to an employee and a catalog skill, then diff it.

    `existing` is the current SkillAssessment rows, used for the diff. `allowed_fields`
    is the actor's writable set — an employee self-assessing may only supply Self and
    Evidence, and anything else is dropped with a `forbidden_field` note rather than
    failing the whole upload.
    """
    by_email, by_code, by_name = _index_employees(employees)
    by_skill_code = {}
    by_skill_slug = {}
    by_skill_name = {}
    for skill in catalog:
        by_skill_code[str(skill.code)] = skill
        by_skill_slug[skill.id.strip().lower()] = skill
        key = (skill.name or "").strip().lower()
        by_skill_name[key] = None if key in by_skill_name else skill

    current = {(a.employee_id, a.skill_id): a for a in existing}
    catalog_by_id = {s.id: s for s in catalog}

    resolved: list[ResolvedRow] = []
    counts = {key: 0 for key in COUNT_KEYS}
    counts["rows_read"] = len(parsed_rows)
    seen: dict[tuple[str, str], ResolvedRow] = {}
    warnings: list[str] = []
    derived_disagreements = 0

    for parsed in parsed_rows:
        row = ResolvedRow(
            row_number=parsed.row_number,
            status="ok",
            raw=parsed.raw,
            assessed_at=parsed.assessed_at.isoformat() if parsed.assessed_at else None,
        )

        if parsed.errors:
            # Carry whatever identified the row through, so the preview can say *whose*
            # row is bad rather than just quoting a row number.
            row.status = "invalid_value"
            row.employee_name = (
                parsed.employee_name or parsed.employee_email or parsed.employee_code
            )
            row.skill_name = parsed.skill_name or parsed.skill_code
            row.messages.extend(parsed.errors)
            resolved.append(row)
            counts["invalid_value"] += 1
            continue

        # --- employee -------------------------------------------------------
        employee = None
        if parsed.employee_email:
            hit = by_email.get(parsed.employee_email.strip().lower(), "missing")
            if hit is None:
                row.status = "ambiguous_employee"
                row.messages.append(f"More than one employee has the email {parsed.employee_email}.")
            elif hit != "missing":
                employee, row.matched_by = hit, "email"
        if employee is None and row.status == "ok" and parsed.employee_code:
            hit = by_code.get(parsed.employee_code.strip().upper(), "missing")
            if hit is None:
                row.status = "ambiguous_employee"
                row.messages.append(f"More than one employee has the ID {parsed.employee_code}.")
            elif hit != "missing":
                employee, row.matched_by = hit, "employee_id"
        if employee is None and row.status == "ok" and parsed.employee_name:
            key = re.sub(r"\s+", " ", parsed.employee_name.strip().lower())
            hit = by_name.get(key, "missing")
            if hit is None:
                row.status = "ambiguous_employee"
                row.messages.append(
                    f"More than one employee is named {parsed.employee_name}. "
                    "Add an Email or Employee ID column to disambiguate."
                )
            elif hit != "missing":
                employee, row.matched_by = hit, "name"

        if row.status == "ambiguous_employee":
            resolved.append(row)
            counts["ambiguous_employee"] += 1
            continue
        if employee is None:
            row.status = "unknown_employee"
            row.employee_name = parsed.employee_name or parsed.employee_email or parsed.employee_code
            row.messages.append("No team member matched this row.")
            resolved.append(row)
            counts["unknown_employee"] += 1
            continue

        row.employee_id, row.employee_name = employee.id, employee.name

        if restrict_to_employee_id and employee.id != restrict_to_employee_id:
            row.status = "forbidden_field"
            row.messages.append("You may only submit an assessment for yourself.")
            resolved.append(row)
            counts["forbidden_field"] += 1
            continue

        # --- skill ----------------------------------------------------------
        skill = None
        if parsed.skill_code:
            key = parsed.skill_code.strip()
            skill = by_skill_code.get(key) or by_skill_slug.get(key.lower())
        if skill is None and parsed.skill_name:
            hit = by_skill_name.get(parsed.skill_name.strip().lower())
            if hit is not None:
                skill = hit
        if skill is None:
            row.status = "unknown_skill"
            row.skill_name = parsed.skill_name or parsed.skill_code
            row.messages.append("This skill is not in the catalog.")
            resolved.append(row)
            counts["unknown_skill"] += 1
            continue

        row.skill_id, row.skill_name = skill.id, skill.name

        # --- field-level authorization --------------------------------------
        values: dict[str, Any] = {}
        dropped: list[str] = []
        for name, value in parsed.values.items():
            if name in allowed_fields:
                values[name] = value
            else:
                dropped.append(name)
        for name in parsed.clears:
            if name in allowed_fields:
                values[name] = None
            else:
                dropped.append(name)
        if dropped:
            row.messages.append(
                "Ignored " + ", ".join(sorted(set(dropped))) + " — not yours to set."
            )

        # --- target_override: only when it differs from the catalog ----------
        # effectiveTarget() is `targetOverride ?? def.targetLevel`, and the export
        # writes a Target on every row. Storing it verbatim would pin an override on
        # every assessment, and editing the catalog target would then silently stop
        # affecting anyone who had ever been imported.
        if "target_override" in values and values["target_override"] is not None:
            if values["target_override"] == catalog_by_id[skill.id].target_level:
                values["target_override"] = None

        if not values:
            row.status = "empty"
            resolved.append(row)
            counts["empty"] += 1
            continue

        # --- duplicate within the file --------------------------------------
        key = (employee.id, skill.id)
        if key in seen:
            earlier = seen[key]
            earlier.status = "duplicate"
            earlier.messages.append(
                f"Superseded by row {parsed.row_number} for the same person and skill."
            )
            earlier.values = {}

        # --- diff against stored ---------------------------------------------
        stored = current.get(key)
        before = {
            name: getattr(stored, name, None) if stored is not None else None
            for name in values
        }
        row.before = before
        row.values = values
        changed = any(before[name] != values[name] for name in values)
        row.status = "ok" if changed else "unchanged"

        # Final/Gap/Priority in the file are recomputed, never stored. They cache as
        # empty in the shipped workbooks, so only compare when actually present.
        if parsed.file_derived:
            expected_final = values.get("reviewer_rating")
            if expected_final is None:
                expected_final = values.get("self_rating")
                if expected_final is None and stored is not None:
                    expected_final = stored.reviewer_rating if stored.reviewer_rating is not None else stored.self_rating
            file_final = parsed.file_derived.get("file_final")
            if file_final and expected_final is not None and str(expected_final) != str(file_final):
                derived_disagreements += 1

        seen[key] = row
        resolved.append(row)

    counts["ok"] = sum(1 for r in resolved if r.status == "ok")
    counts["unchanged"] = sum(1 for r in resolved if r.status == "unchanged")
    counts["duplicate"] = sum(1 for r in resolved if r.status == "duplicate")

    if derived_disagreements:
        warnings.append(
            f"{derived_disagreements} row(s) have a Final rating that disagrees with their "
            "Self/Reviewer values. Final, Gap and Priority are always recalculated, so the "
            "file's versions were ignored."
        )

    return ImportPreviewResult(rows=resolved, counts=counts, warnings=warnings)
